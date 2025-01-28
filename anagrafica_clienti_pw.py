# script creato da Pastore Pietro come richiesto nel PW Pegaso
# script per generare un file Excel con un'anagrafica di clienti
# con nome, cognome, email, telefono e due hash delle password
# SHA256 e SHA512

# Importiamo le librerie necessarie
import random          # Serve per generare numeri e scelte casuali
import string         # Serve per accedere a lettere e caratteri speciali
import hashlib        # Serve per crittografare le password
import os             # Serve per operazioni sul sistema operativo (creazione cartelle)
from openpyxl import Workbook    # Serve per creare e gestire file Excel
import names          # Serve per generare nomi casuali realistici
import faker          # Serve per generare vari tipi di dati casuali realistici

def generate_password_hashes():
    """Questa funzione genera una password casuale e ne restituisce gli hash SHA256 e SHA512"""
    # Definisce i caratteri possibili per la password (lettere, numeri e simboli)
    chars = string.ascii_letters + string.digits + string.punctuation
    # Crea una password randomdi 12 caratteri scegliendo casualmente dai caratteri possibili
    # Questo simula una ipotetica password inserita da un utente in fase di registrazione
    password = ''.join(random.choice(chars) for _ in range(12))
    
    # Genera entrambi gli hash della password
    # SHA256 - hash a 256 bit (64 caratteri esadecimali)
    hash_256 = hashlib.sha256(password.encode()).hexdigest()
    # SHA512 - hash a 512 bit (128 caratteri esadecimali) - più sicuro
    hash_512 = hashlib.sha512(password.encode()).hexdigest()
    
    return hash_256, hash_512

def generate_phone():
    """Questa funzione genera un numero di telefono italiano casuale"""
    # Lista dei prefissi mobili italiani più comuni
    prefissi = ['320', '330', '340', '350', '360', '370', '380', '390']
    # Crea il numero concatenando +39 (prefisso Italia), un prefisso mobile e 7 numeri casuali
    return f"+39{random.choice(prefissi)}{random.randint(1000000,9999999)}"

def main():
    # Chiede all'utente quanti record generare con gestione degli errori
    while True:
        try:
            # Prova a convertire l'input in numero
            num_users = int(input("Inserisci il numero di utenti da generare: "))
            # Verifica che sia un numero positivo
            if num_users > 0:
                break
            print("Per favore inserisci un numero positivo")
        except ValueError:
            print("Per favore inserisci un numero valido")
    
    # Definisce il percorso della cartella dove salvare il file
    folder_path = "C:/Project_Work_Pegaso"
    # Se la cartella non esiste, la crea
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    # Inizializza il generatore di dati faker per l'Italia
    # Questo ci permetterà di generare dati più realistici per il contesto italiano
    fake = faker.Faker('it_IT')
    
    # Crea un nuovo file Excel
    wb = Workbook()
    ws = wb.active  # Prende il foglio attivo
    ws.title = "Anagrafica Clienti"  # Imposta il titolo del foglio
    
    # Definisce le intestazioni delle colonne
    headers = ['Nome', 'Cognome', 'Email', 'Telefono', 'Password Hash (SHA256)', 'Password Hash (SHA512)']
    # Inserisce le intestazioni nella prima riga
    # enumerate(headers, 1) fa partire l'indice da 1 invece che da 0
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Genera i dati per ogni utente
    for row in range(2, num_users + 2):  # Parte da riga 2 (dopo le intestazioni)
        # Genera un nome casuale
        nome = names.get_first_name()
        # Genera un cognome casuale
        cognome = names.get_last_name()
        # Crea una email basata su nome e cognome (tutto minuscolo)
        email = f"{nome.lower()}.{cognome.lower()}@{fake.free_email_domain()}"
        # Genera un numero di telefono italiano
        telefono = generate_phone()
        # Genera gli hash delle password
        hash_256, hash_512 = generate_password_hashes()
        
        # Inserisce i dati nelle celle del foglio Excel
        ws.cell(row=row, column=1, value=nome)
        ws.cell(row=row, column=2, value=cognome)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=telefono)
        ws.cell(row=row, column=5, value=hash_256)
        ws.cell(row=row, column=6, value=hash_512)
    
    # Crea il percorso completo del file
    file_path = os.path.join(folder_path, "Anagrafica_clienti_Pw_Pegaso.xlsx")
    # Salva il file Excel
    wb.save(file_path)
    # Stampa un messaggio di conferma con il percorso del file
    print(f"\nFile generato con successo in: {file_path}")

# Questo è il punto di ingresso del programma
# Il codice viene eseguito solo se il file viene eseguito direttamente
if __name__ == "__main__":
    main()  # Chiama la funzione principale